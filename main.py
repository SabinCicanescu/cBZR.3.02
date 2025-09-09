import pandas as pd
from openpyxl import load_workbook
from tkinter import Tk, filedialog, simpledialog

# Hide root window
root = Tk()
root.withdraw()

# Select source Excel files
source_files = filedialog.askopenfilenames(title="Select Excel Files", filetypes=[("Excel files", "*.xlsx")])
if not source_files:
    print("No files selected.")
    exit()

# Select master file
master_file = filedialog.askopenfilename(title="Select Master Excel File", filetypes=[("Excel files", "*.xlsx")])
if not master_file:
    print("No master file selected.")
    exit()

# Ask for sheet name
sheet_name = simpledialog.askstring("Sheet Name", "Enter the sheet name to append data to:")
if not sheet_name:
    print("No sheet name provided.")
    exit()

# Load master workbook and find last row
book = load_workbook(master_file)
if sheet_name not in book.sheetnames:
    book.create_sheet(sheet_name)
    last_row = 0
else:
    last_row = book[sheet_name].max_row

writer = pd.ExcelWriter(master_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
writer.book = book

# Loop through selected files and append data
for file in source_files:
    df = pd.read_excel(file, sheet_name=sheet_name)
    df.to_excel(writer, sheet_name=sheet_name, startrow=last_row, index=False, header=False)
    last_row += len(df)

# Save changes
writer.close()
print("Data successfully appended to master file.")

